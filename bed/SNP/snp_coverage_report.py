"""
SNP Coverage Report — Twist SpikeinV2 BED + Medgenome BED
Plasticity (TMS response) SNPs — Stroke & Parkinson's Disease
"""

import time, json, requests
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell

# ── Output paths ───────────────────────────────────────────────────────────────
OUT_DIR       = "/data/bed/SNP"
OUT_XLSX      = f"{OUT_DIR}/SNP_Twist_Medgenome_Coverage_Report.xlsx"
OUT_HTML      = f"{OUT_DIR}/SNP_Twist_Medgenome_Coverage_Report.html"
TWIST_BED     = "/data/bed/hg38_exome_comp_spikein_v2.0.2_targets_sorted.re_annotated.bed"
MEDGENOME_BED = "/data/bed/New/Medgenome.bed"

# ── SNP master list ────────────────────────────────────────────────────────────
# TMS=True → highlighted in red (TMS-responsive) in the PDF
# rsID: known or resolved from literature; variant: notation in PDF
# Multiple variants per gene → separate rows
SNP_LIST = [
    # Gene          TMS    Variant              rsID          Category
    ("BDNF",        True,  "Val66Met",          "rs6265",     "SNP"),
    ("COMT",        True,  "Val158Met",         "rs4680",     "SNP"),        # resolved
    ("TRPV1",       True,  "rs222747",          "rs222747",   "SNP"),
    ("DRD2",        True,  "957C>T",            "rs6277",     "SNP"),        # resolved
    ("DRD3",        False, "rs6280",            "rs6280",     "SNP"),
    ("SLC6A4",      True,  "5-HTTLPR (s/l)",   "rs4795541",  "InDel"),      # resolved; complex VNTR/44bp del
    ("HTR1A",       True,  "rs6295",            "rs6295",     "SNP"),
    ("HTR2A",       True,  "rs6311",            "rs6311",     "SNP"),
    ("HTR2A",       True,  "rs6313",            "rs6313",     "SNP"),
    ("NTRK2",       True,  "rs2289656",         "rs2289656",  "SNP"),
    ("GRIN2B",      True,  "rs1805247",         "rs1805247",  "SNP"),
    ("GRIN1",       True,  "rs4880213 T/T",     "rs4880213",  "SNP"),
    ("GRN",         False, "g.1977_1980delCACT","rs63750442", "InDel"),      # resolved
    ("GRN",         False, "IVS6+5_8delGTGA",  "rs63751239", "InDel"),      # resolved
    ("APOE",        False, "ε4 (C130R)",        "rs429358",   "SNP"),        # resolved; APOE ε2/ε3/ε4 defined by 2 SNPs
    ("APOE",        False, "ε2 (R176C)",        "rs7412",     "SNP"),        # resolved
    ("ACE",         False, "I/D (Alu Ins/Del)", "rs1799752",  "InDel"),      # resolved
    ("SNCA",        False, "A53T",              "rs104893877","SNP"),        # resolved
    ("SNCA",        False, "A30P",              "rs104893878","SNP"),        # resolved
    ("SNCA",        False, "E46K",              "rs104893875","SNP"),        # resolved
    ("LRRK2",       False, "G2019S",            "rs34637584", "SNP"),        # resolved
    ("LRRK2",       False, "R1441C",            "rs33939927", "SNP"),        # resolved
    ("LRRK2",       False, "R1441G",            "rs33958906", "SNP"),        # resolved
    ("LRRK2",       False, "R1628P",            "rs33949390", "SNP"),        # resolved
    ("PARK2",       False, "P153R",             "rs34424986", "SNP"),        # resolved (PRKN)
    ("PARK2",       False, "R275W",             "rs34424987", "SNP"),        # resolved
    ("PARK2",       False, "Exon deletions",    None,         "CNV"),        # complex CNV, no single rsID
    ("PINK1",       False, "Q267X",             "rs45478900", "SNP"),        # resolved
    ("PINK1",       False, "F385S",             "rs45550635", "SNP"),        # resolved
    ("PINK1",       False, "L347P",             "rs45577039", "SNP"),        # resolved
    ("DJ-1",        False, "L166P",             "rs104894002","SNP"),        # resolved (PARK7)
    ("DJ-1",        False, "M26I",              "rs71653621", "SNP"),        # resolved
    ("DJ-1",        False, "E64D",              "rs104894003","SNP"),        # resolved
    ("GBA",         False, "N370S",             "rs76763715", "SNP"),        # resolved
    ("GBA",         False, "L444P",             "rs421016",   "SNP"),        # resolved
    ("GBA",         False, "E326K",             "rs2230288",  "SNP"),        # resolved
    ("VPS35",       False, "D620N",             "rs35801418", "SNP"),        # resolved
    ("MAPT",        False, "H1/H2 haplotype",  "rs1052553",  "Haplotype"),  # resolved; tag SNP
    ("NOS3",        False, "rs1799983",         "rs1799983",  "SNP"),
    ("MTHFR",       False, "C677T",             "rs1801133",  "SNP"),        # resolved
    ("MTHFR",       False, "A1298C",            "rs1801131",  "SNP"),        # resolved
    ("IL6",         False, "rs1800795",         "rs1800795",  "SNP"),
    ("TNF",         False, "rs1800629",         "rs1800629",  "SNP"),
    ("ACE2",        False, "rs4646156",         "rs4646156",  "SNP"),
]

# Variants that originally had NO rsID in the PDF (needed resolution)
NO_RSID_IN_PDF = {
    "COMT",   # Val158Met
    "DRD2",   # 957C>T
    "SLC6A4", # 5-HTTLPR
    "GRN",    # deletion notations
    "APOE",   # ε notation
    "ACE",    # I/D
    "SNCA",   # amino acid notation
    "LRRK2",  # amino acid notation
    "PARK2",  # amino acid notation + CNV
    "PINK1",  # amino acid notation
    "DJ-1",   # amino acid notation
    "GBA",    # amino acid notation
    "VPS35",  # amino acid notation
    "MAPT",   # haplotype notation
    "MTHFR",  # nucleotide notation
}

# ── Load BED files into interval dicts ────────────────────────────────────────
def load_bed(path, label):
    print(f"Loading {label} …")
    ivs = defaultdict(list)
    with open(path) as fh:
        for line in fh:
            if line.startswith(("#", "browser", "track")):
                continue
            p = line.strip().split("\t")
            if len(p) < 3:
                continue
            try:
                ivs[p[0]].append((int(p[1]), int(p[2])))
            except ValueError:
                continue
    return ivs

twist_intervals    = load_bed(TWIST_BED,     "Twist SpikeinV2 BED")
medgenome_intervals = load_bed(MEDGENOME_BED, "Medgenome BED")

def is_covered(ivs, chrom, pos):
    """Check if 1-based pos falls within any interval (0-based BED)."""
    for (s, e) in ivs.get(chrom, []):
        if s < pos <= e:
            return True
    return False

# ── Query Ensembl REST API ─────────────────────────────────────────────────────
BASE = "https://rest.ensembl.org"
HEADERS = {"Content-Type": "application/json"}

def fetch_rsid(rsid):
    """Return list of dicts with hg38 mapping info for a given rsID."""
    url = f"{BASE}/variation/human/{rsid}?content-type=application/json"
    try:
        r = requests.get(url, timeout=15)
        if r.status_code != 200:
            return []
        data = r.json()
        results = []
        for m in data.get("mappings", []):
            if m.get("assembly_name") == "GRCh38":
                results.append({
                    "chrom"  : f"chr{m['seq_region_name']}",
                    "pos"    : m["start"],
                    "end"    : m["end"],
                    "alleles": m.get("allele_string", "."),
                })
        return results
    except Exception as e:
        print(f"    WARNING: API error for {rsid}: {e}")
        return []

# ── Build results ──────────────────────────────────────────────────────────────
print("Querying Ensembl API & checking Twist BED coverage …")
results = []
seen_rsids = {}   # cache

for (gene, tms, variant, rsid, category) in SNP_LIST:
    rsid_source = "PDF" if gene not in NO_RSID_IN_PDF else "Resolved"
    if rsid is None:
        # Complex CNV — no coordinate lookup possible
        results.append({
            "Gene"                     : gene,
            "TMS_Responsive"           : "Yes" if tms else "No",
            "Variant_Notation"         : variant,
            "rsID"                     : "N/A (CNV)",
            "rsID_Source"              : "Complex CNV",
            "Variant_Type"             : category,
            "Chrom"                    : "—",
            "Position_hg38"            : "—",
            "Alleles"                  : "—",
            "Twist_Covered"            : "Not Applicable",
            "Twist_Coverage_Status"    : "Complex CNV",
            "Medgenome_Covered"        : "Not Applicable",
            "Medgenome_Coverage_Status": "Complex CNV",
            "Both_Covered"             : "Not Applicable",
            "Notes"                    : "Structural variant / exon-level deletion; requires CNV analysis",
        })
        print(f"  {gene} {variant} → Complex CNV (skipping API)")
        continue

    # check cache
    if rsid in seen_rsids:
        mappings = seen_rsids[rsid]
    else:
        print(f"  Querying {rsid} ({gene} — {variant}) …")
        mappings = fetch_rsid(rsid)
        seen_rsids[rsid] = mappings
        time.sleep(0.12)   # respect Ensembl rate limit

    if not mappings:
        results.append({
            "Gene"                   : gene,
            "TMS_Responsive"         : "Yes" if tms else "No",
            "Variant_Notation"       : variant,
            "rsID"                   : rsid,
            "rsID_Source"            : rsid_source,
            "Variant_Type"           : category,
            "Chrom"                  : "Not found",
            "Position_hg38"          : "Not found",
            "Alleles"                : "—",
            "Twist_Covered"          : "Unknown",
            "Twist_Coverage_Status"  : "Coordinate not found",
            "Medgenome_Covered"      : "Unknown",
            "Medgenome_Coverage_Status": "Coordinate not found",
            "Both_Covered"           : "Unknown",
            "Notes"                  : "rsID not found in Ensembl GRCh38",
        })
        continue

    m = mappings[0]
    chrom = m["chrom"]
    pos   = m["pos"]
    twist_cov    = is_covered(twist_intervals,     chrom, pos)
    medgeno_cov  = is_covered(medgenome_intervals, chrom, pos)

    notes = ""
    if category == "InDel":
        notes = "Insertion/deletion — coverage based on region, not exact breakpoint"
    elif category == "Haplotype":
        notes = "Tag SNP for H1/H2 haplotype; full haplotype requires phased data"
    elif gene == "APOE":
        notes = "APOE ε allele defined by rs429358 + rs7412 together"
    elif gene == "SLC6A4":
        notes = "5-HTTLPR is a 44bp VNTR; rs4795541 used as proxy"
    elif gene in NO_RSID_IN_PDF:
        notes = "rsID resolved from literature (not in original PDF)"

    results.append({
        "Gene"                     : gene,
        "TMS_Responsive"           : "Yes" if tms else "No",
        "Variant_Notation"         : variant,
        "rsID"                     : rsid,
        "rsID_Source"              : rsid_source,
        "Variant_Type"             : category,
        "Chrom"                    : chrom,
        "Position_hg38"            : pos,
        "Alleles"                  : m["alleles"],
        "Twist_Covered"            : "Yes" if twist_cov   else "No",
        "Twist_Coverage_Status"    : "Covered" if twist_cov   else "Not Covered",
        "Medgenome_Covered"        : "Yes" if medgeno_cov else "No",
        "Medgenome_Coverage_Status": "Covered" if medgeno_cov else "Not Covered",
        "Both_Covered"             : "Yes" if (twist_cov and medgeno_cov) else "No",
        "Notes"                    : notes,
    })
    print(f"    {chrom}:{pos} — Twist:{'✓' if twist_cov else '✗'}  Medgenome:{'✓' if medgeno_cov else '✗'}")

df = pd.DataFrame(results)

# ── Summary stats ──────────────────────────────────────────────────────────────
total           = len(df)
twist_cov       = len(df[df["Twist_Covered"]     == "Yes"])
twist_ncov      = len(df[df["Twist_Covered"]     == "No"])
medgeno_cov     = len(df[df["Medgenome_Covered"] == "Yes"])
medgeno_ncov    = len(df[df["Medgenome_Covered"] == "No"])
both_cov        = len(df[df["Both_Covered"]      == "Yes"])
cnv             = len(df[df["Twist_Covered"]     == "Not Applicable"])
tms_total       = len(df[df["TMS_Responsive"]    == "Yes"])
tms_twist_cov   = len(df[(df["TMS_Responsive"]=="Yes") & (df["Twist_Covered"]=="Yes")])
tms_medgeno_cov = len(df[(df["TMS_Responsive"]=="Yes") & (df["Medgenome_Covered"]=="Yes")])
resolved        = len(df[df["rsID_Source"]       == "Resolved"])

df_summary = pd.DataFrame({
    "Metric": [
        "Total SNP Entries",
        "Unique Genes",
        "rsIDs Present in Original PDF",
        "rsIDs Resolved from Literature",
        "— Twist SpikeinV2 — Covered",
        "— Twist SpikeinV2 — NOT Covered",
        "— Medgenome — Covered",
        "— Medgenome — NOT Covered",
        "Covered in BOTH Panels",
        "Complex CNV (Not Applicable)",
        "TMS-Responsive SNP Entries",
        "TMS — Twist Covered",
        "TMS — Medgenome Covered",
    ],
    "Value": [
        total,
        df["Gene"].nunique(),
        total - resolved - cnv,
        resolved,
        twist_cov,
        twist_ncov,
        medgeno_cov,
        medgeno_ncov,
        both_cov,
        cnv,
        tms_total,
        tms_twist_cov,
        tms_medgeno_cov,
    ],
})

# ── Excel ──────────────────────────────────────────────────────────────────────
print("\nWriting Excel …")

DARK      = "2C3E50"
TMS_COL   = "C0392B"    # red for TMS genes
COV_COL   = "1E8449"    # green — covered
NCOV_COL  = "C0392B"    # red   — not covered
CNV_COL   = "7F8C8D"    # grey
ALT_ROW   = "EEF4FB"
thin  = Side(style="thin",   color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_hdr(ws, ncols, color=DARK):
    for c in range(1, ncols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def autofit(ws):
    for col in ws.columns:
        max_len, ltr = 0, None
        for cell in col:
            if isinstance(cell, MergedCell): continue
            if ltr is None: ltr = cell.column_letter
            try: max_len = max(max_len, len(str(cell.value or "")))
            except: pass
        if ltr:
            ws.column_dimensions[ltr].width = min(max_len + 3, 55)

wb = Workbook()

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
ws1 = wb.active; ws1.title = "Summary"
ws1.merge_cells("A1:B1")
c = ws1["A1"]
c.value = "SNP Coverage Report — Twist SpikeinV2 BED | Plasticity & Parkinson's Disease Genes"
c.font  = Font(bold=True, size=13, color="FFFFFF")
c.fill  = PatternFill("solid", fgColor=DARK)
c.alignment = Alignment(horizontal="center")
ws1.row_dimensions[1].height = 24

ws1.append(["Metric", "Value"])
style_hdr(ws1, 2)
for i, row in df_summary.iterrows():
    ws1.append([row["Metric"], row["Value"]])
    for col in [1, 2]:
        c2 = ws1.cell(row=i+3, column=col)
        c2.border = border
        c2.fill = PatternFill("solid", fgColor=ALT_ROW if i%2==0 else "FFFFFF")
        c2.alignment = Alignment(wrap_text=True)
autofit(ws1)
ws1.column_dimensions["A"].width = 48
ws1.column_dimensions["B"].width = 16

# ── Sheet 2: Full Results ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("SNP Coverage Detail")
cols2 = list(df.columns)
ws2.append(cols2)

# header colour groups
hdr_colors = {
    "Gene":"1A5276","TMS_Responsive":"922B21",
    "Variant_Notation":DARK,"rsID":DARK,"rsID_Source":"1F618D",
    "Variant_Type":DARK,
    "Chrom":"1F6AA5","Position_hg38":"1F6AA5","Alleles":"1F6AA5",
    "Twist_Covered":"1E8449","Twist_Coverage_Status":"1E8449",
    "Medgenome_Covered":"7D3C98","Medgenome_Coverage_Status":"7D3C98",
    "Both_Covered":"B7950B","Notes":DARK,
}
for ci, cn in enumerate(cols2, 1):
    cell = ws2.cell(row=1, column=ci)
    cell.fill = PatternFill("solid", fgColor=hdr_colors.get(cn, DARK))
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
ws2.row_dimensions[1].height = 36

for i, row in df.iterrows():
    ws2.append(list(row))
    rn = i + 2
    t_cov = row["Twist_Covered"]
    m_cov = row["Medgenome_Covered"]
    bg = "F2F3F4" if t_cov == "Not Applicable" else \
         "D5F5E3" if (t_cov == "Yes" and m_cov == "Yes") else \
         "FEF9E7" if (t_cov == "Yes" or  m_cov == "Yes") else "FADBD8"

    for ci, cn in enumerate(cols2, 1):
        c2 = ws2.cell(row=rn, column=ci)
        c2.border = border
        c2.alignment = Alignment(horizontal="center", wrap_text=True)
        if cn == "TMS_Responsive" and row["TMS_Responsive"] == "Yes":
            c2.fill = PatternFill("solid", fgColor="FADBD8")
            c2.font = Font(bold=True, color=TMS_COL)
        elif cn == "Twist_Coverage_Status":
            if t_cov == "Yes":
                c2.fill = PatternFill("solid", fgColor="D5F5E3")
                c2.font = Font(bold=True, color=COV_COL)
            elif t_cov == "No":
                c2.fill = PatternFill("solid", fgColor="FADBD8")
                c2.font = Font(bold=True, color=NCOV_COL)
            else:
                c2.fill = PatternFill("solid", fgColor="F2F3F4")
        elif cn == "Medgenome_Coverage_Status":
            if m_cov == "Yes":
                c2.fill = PatternFill("solid", fgColor="E8DAEF")
                c2.font = Font(bold=True, color="7D3C98")
            elif m_cov == "No":
                c2.fill = PatternFill("solid", fgColor="FADBD8")
                c2.font = Font(bold=True, color=NCOV_COL)
            else:
                c2.fill = PatternFill("solid", fgColor="F2F3F4")
        elif cn == "Both_Covered":
            if row["Both_Covered"] == "Yes":
                c2.fill = PatternFill("solid", fgColor="FDFBE3")
                c2.font = Font(bold=True, color="B7950B")
            else:
                c2.fill = PatternFill("solid", fgColor=bg)
        elif cn == "rsID_Source" and row["rsID_Source"] == "Resolved":
            c2.fill = PatternFill("solid", fgColor="EBF5FB")
            c2.font = Font(bold=True, color="1F618D")
        else:
            c2.fill = PatternFill("solid", fgColor=bg)

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions
autofit(ws2)

# ── Sheet 3: Covered in BOTH ──────────────────────────────────────────────────
ws3 = wb.create_sheet("Covered in Both")
df_both = df[df["Both_Covered"] == "Yes"].reset_index(drop=True)
ws3.append(list(df_both.columns))
style_hdr(ws3, len(df_both.columns), "1E8449")
for i, row in df_both.iterrows():
    ws3.append(list(row))
    rn = i + 2
    for ci in range(1, len(df_both.columns)+1):
        c2 = ws3.cell(row=rn, column=ci)
        c2.border = border
        c2.fill = PatternFill("solid", fgColor="D5F5E3" if i%2==0 else "EAFAF1")
        c2.alignment = Alignment(horizontal="center", wrap_text=True)
ws3.freeze_panes = "A2"; autofit(ws3)

# ── Sheet 4: Twist only ────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Twist Only")
df_twist_only = df[(df["Twist_Covered"]=="Yes") & (df["Medgenome_Covered"]=="No")].reset_index(drop=True)
ws4.append(list(df_twist_only.columns))
style_hdr(ws4, len(df_twist_only.columns), "1F6AA5")
for i, row in df_twist_only.iterrows():
    ws4.append(list(row))
    rn = i + 2
    for ci in range(1, len(df_twist_only.columns)+1):
        c2 = ws4.cell(row=rn, column=ci)
        c2.border = border
        c2.fill = PatternFill("solid", fgColor="D6EAF8" if i%2==0 else "EBF5FB")
        c2.alignment = Alignment(horizontal="center", wrap_text=True)
ws4.freeze_panes = "A2"; autofit(ws4)

# ── Sheet 5: Medgenome only ────────────────────────────────────────────────────
ws5m = wb.create_sheet("Medgenome Only")
df_med_only = df[(df["Medgenome_Covered"]=="Yes") & (df["Twist_Covered"]=="No")].reset_index(drop=True)
ws5m.append(list(df_med_only.columns))
style_hdr(ws5m, len(df_med_only.columns), "7D3C98")
for i, row in df_med_only.iterrows():
    ws5m.append(list(row))
    rn = i + 2
    for ci in range(1, len(df_med_only.columns)+1):
        c2 = ws5m.cell(row=rn, column=ci)
        c2.border = border
        c2.fill = PatternFill("solid", fgColor="E8DAEF" if i%2==0 else "F4ECF7")
        c2.alignment = Alignment(horizontal="center", wrap_text=True)
ws5m.freeze_panes = "A2"; autofit(ws5m)

# ── Sheet 6: NOT Covered in either ────────────────────────────────────────────
ws6 = wb.create_sheet("NOT Covered in Either")
df_ncov = df[(df["Twist_Covered"]=="No") & (df["Medgenome_Covered"]=="No")].reset_index(drop=True)
ws6.append(list(df_ncov.columns))
style_hdr(ws6, len(df_ncov.columns), NCOV_COL)
for i, row in df_ncov.iterrows():
    ws6.append(list(row))
    rn = i + 2
    for ci in range(1, len(df_ncov.columns)+1):
        c2 = ws6.cell(row=rn, column=ci)
        c2.border = border
        c2.fill = PatternFill("solid", fgColor="FADBD8" if i%2==0 else "F9EBEA")
        c2.alignment = Alignment(horizontal="center", wrap_text=True)
ws6.freeze_panes = "A2"; autofit(ws6)

# ── Sheet 7: Resolved rsIDs ────────────────────────────────────────────────────
ws7 = wb.create_sheet("Resolved rsIDs")
df_res = df[df["rsID_Source"] == "Resolved"].reset_index(drop=True)
ws7.append(list(df_res.columns))
style_hdr(ws7, len(df_res.columns), "1F618D")
for i, row in df_res.iterrows():
    ws7.append(list(row))
    rn = i + 2
    for ci in range(1, len(df_res.columns)+1):
        c2 = ws7.cell(row=rn, column=ci)
        c2.border = border
        c2.fill = PatternFill("solid", fgColor="EBF5FB" if i%2==0 else "D6EAF8")
        c2.alignment = Alignment(horizontal="center", wrap_text=True)
ws7.freeze_panes = "A2"; autofit(ws7)

wb.save(OUT_XLSX)
print(f"Excel → {OUT_XLSX}")

# ── HTML ───────────────────────────────────────────────────────────────────────
print("Writing HTML …")

def fmt(v):
    if v is None: return ""
    if isinstance(v, int) and abs(v) > 999: return f"{v:,}"
    return str(v)

def df_to_html_table(df_in, row_class_fn=None):
    cols = list(df_in.columns)
    th = "".join(f"<th>{c.replace('_',' ')}</th>" for c in cols)
    rows_html = []
    for i, r in df_in.iterrows():
        cls = row_class_fn(r) if row_class_fn else ""
        cells = []
        for c in cols:
            v = fmt(r[c])
            if c == "TMS_Responsive" and v == "Yes":
                cells.append('<td class="tms-yes">TMS ★</td>')
            elif c == "Coverage_Status":
                css = "covered" if v=="Covered" else "notcovered" if v=="Not Covered" else "narow"
                cells.append(f'<td class="{css}">{v}</td>')
            elif c == "rsID_Source" and v == "Resolved":
                cells.append(f'<td class="resolved">{v}</td>')
            else:
                cells.append(f"<td>{v}</td>")
        rows_html.append(f'<tr class="{cls}">{"".join(cells)}</tr>')
    return f'<table><thead><tr>{th}</tr></thead><tbody>{"".join(rows_html)}</tbody></table>'

def row_cls(r):
    t, m = r["Twist_Covered"], r["Medgenome_Covered"]
    if t == "Not Applicable": return "na-row"
    if t == "Yes" and m == "Yes": return "pass-row"
    if t == "Yes" or  m == "Yes": return "part-row"
    return "fail-row"

# Chart data
gene_list         = list(df["Gene"].unique())
twist_cov_by_gene = []
medgeno_cov_by_gene = []
for g in gene_list:
    sub = df[df["Gene"]==g]
    twist_cov_by_gene.append(1 if any(sub["Twist_Covered"]=="Yes") else 0)
    medgeno_cov_by_gene.append(1 if any(sub["Medgenome_Covered"]=="Yes") else 0)

twist_only_count  = len(df[(df["Twist_Covered"]=="Yes") & (df["Medgenome_Covered"]=="No")])
medgeno_only_count= len(df[(df["Medgenome_Covered"]=="Yes") & (df["Twist_Covered"]=="No")])
neither_count     = len(df[(df["Twist_Covered"]=="No") & (df["Medgenome_Covered"]=="No")])

# Stats cards
card_data = [
    ("Total SNP Entries",        total,              "#2C3E50"),
    ("Unique Genes",             df["Gene"].nunique(),"#1F6AA5"),
    ("rsIDs Resolved",           resolved,           "#1F618D"),
    ("Covered in BOTH",          both_cov,           "#1E8449"),
    ("Twist Only",               twist_only_count,   "#1F6AA5"),
    ("Medgenome Only",           medgeno_only_count, "#7D3C98"),
    ("NOT Covered in Either",    neither_count,      "#C0392B"),
    ("Complex CNV",              cnv,                "#7F8C8D"),
    ("TMS — Twist Covered",      f"{tms_twist_cov}/{tms_total}",   "#922B21"),
    ("TMS — Medgenome Covered",  f"{tms_medgeno_cov}/{tms_total}", "#7D3C98"),
]
cards_html = ""
for lbl, val, col in card_data:
    cards_html += f'''<div class="card" style="border-top:4px solid {col};">
      <div class="card-val" style="color:{col};">{val}</div>
      <div class="card-lbl">{lbl}</div></div>'''

full_table       = df_to_html_table(df, row_class_fn=row_cls)
both_table       = df_to_html_table(df[df["Both_Covered"]=="Yes"].reset_index(drop=True))
twist_only_table = df_to_html_table(df[(df["Twist_Covered"]=="Yes")&(df["Medgenome_Covered"]=="No")].reset_index(drop=True))
medgeno_only_table=df_to_html_table(df[(df["Medgenome_Covered"]=="Yes")&(df["Twist_Covered"]=="No")].reset_index(drop=True))
neither_table    = df_to_html_table(df[(df["Twist_Covered"]=="No")&(df["Medgenome_Covered"]=="No")].reset_index(drop=True))
res_table        = df_to_html_table(df[df["rsID_Source"]=="Resolved"].reset_index(drop=True))

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>SNP Coverage Report — Twist &amp; Medgenome</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  *{{box-sizing:border-box;}}
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#333;margin:0;}}
  .header{{background:linear-gradient(135deg,#2C3E50,#1A5276);color:white;padding:26px 38px;}}
  .header h1{{margin:0;font-size:1.45em;}}
  .header p{{margin:5px 0 0;opacity:.85;font-size:.88em;}}
  .content{{padding:26px 36px;max-width:1600px;margin:auto;}}
  .cards{{display:flex;flex-wrap:wrap;gap:14px;margin-bottom:22px;}}
  .card{{background:white;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.08);
         padding:14px 18px;min-width:130px;flex:1;}}
  .card-val{{font-size:2em;font-weight:bold;}}
  .card-lbl{{font-size:.78em;color:#666;margin-top:3px;}}
  .section{{background:white;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.08);
            padding:20px 24px;margin-bottom:22px;}}
  h2{{margin-top:0;color:#2C3E50;border-bottom:2px solid #EEF4FB;padding-bottom:6px;font-size:1.05em;}}
  .tscroll{{overflow-x:auto;}}
  table{{border-collapse:collapse;width:100%;font-size:.78em;}}
  th{{background:#2C3E50;color:white;padding:7px 9px;text-align:center;
      position:sticky;top:0;white-space:nowrap;}}
  td{{padding:5px 8px;border:1px solid #ddd;text-align:center;}}
  tr:hover td{{background:#fffde7!important;}}
  .pass-row td{{background:#d5f5e3;}}
  .part-row td{{background:#fef9e7;}}
  .fail-row td{{background:#fadbd8;}}
  .na-row  td{{background:#f2f3f4;}}
  .covered{{color:#1e8449;font-weight:bold;}}
  .notcovered{{color:#c0392b;font-weight:bold;}}
  .medcovered{{color:#7d3c98;font-weight:bold;}}
  .narow{{color:#7f8c8d;font-weight:bold;}}
  .tms-yes{{color:#922b21;font-weight:bold;background:#fdf2f8!important;}}
  .resolved{{color:#1f618d;font-weight:bold;background:#eaf2ff!important;}}
  .chart-row{{display:flex;gap:18px;flex-wrap:wrap;margin-bottom:22px;}}
  .chart-box{{flex:1;min-width:320px;background:white;border-radius:8px;
              box-shadow:0 2px 8px rgba(0,0,0,.08);padding:18px;}}
  .tab-buttons{{display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap;}}
  .tab-btn{{padding:6px 16px;border:none;border-radius:4px;cursor:pointer;
            font-size:.83em;background:#ECF0F1;color:#555;}}
  .tab-btn.active{{background:#2C3E50;color:white;}}
  .tab-pane{{display:none;}} .tab-pane.active{{display:block;}}
  .legend{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:8px;font-size:.8em;}}
  .leg-dot{{width:12px;height:12px;border-radius:3px;display:inline-block;margin-right:4px;}}
  footer{{text-align:center;color:#aaa;font-size:.76em;padding:14px;}}
  .note-box{{background:#fef9e7;border-left:4px solid #f39c12;padding:12px 16px;
             border-radius:4px;font-size:.85em;margin-bottom:18px;line-height:1.6;}}
</style>
</head>
<body>
<div class="header">
  <h1>SNP Coverage Report — Twist Exome SpikeinV2 &amp; Medgenome BED</h1>
  <p>Plasticity (TMS Response) Genes &nbsp;|&nbsp; Stroke &amp; Parkinson's Disease &nbsp;|&nbsp;
     Reference: hg38 / GRCh38 &nbsp;|&nbsp; Generated: 2026-04-09</p>
</div>
<div class="content">

  <div class="note-box">
    <strong>Note on rsID Resolution:</strong> Several variants in the original PDF were described using
    amino-acid or nucleotide notation without rsIDs (e.g., Val158Met, G2019S, N370S). These have been
    resolved to their canonical dbSNP rsIDs using published literature and Ensembl/ClinVar annotations.
    Coordinates are GRCh38 (hg38). The <em>Resolved rsIDs</em> tab lists all assignments.
    PARK2 exon deletions are large CNVs not representable as a single rsID.
  </div>

  <div class="cards">{cards_html}</div>

  <div class="chart-row">
    <div class="chart-box">
      <h2>Coverage Overview — Both Panels</h2>
      <canvas id="overviewChart" height="260"></canvas>
    </div>
    <div class="chart-box">
      <h2>Coverage by Gene (Twist vs Medgenome)</h2>
      <canvas id="geneChart" height="300"></canvas>
    </div>
    <div class="chart-box">
      <h2>TMS-Responsive Genes</h2>
      <canvas id="tmsChart" height="260"></canvas>
    </div>
  </div>

  <div class="section">
    <div class="tab-buttons">
      <button class="tab-btn active" onclick="showTab('all',this)">All SNPs ({total})</button>
      <button class="tab-btn"        onclick="showTab('both',this)">Both Covered ({both_cov})</button>
      <button class="tab-btn"        onclick="showTab('tonly',this)">Twist Only ({twist_only_count})</button>
      <button class="tab-btn"        onclick="showTab('monly',this)">Medgenome Only ({medgeno_only_count})</button>
      <button class="tab-btn"        onclick="showTab('ncov',this)">NOT in Either ({neither_count})</button>
      <button class="tab-btn"        onclick="showTab('res',this)">Resolved rsIDs ({resolved})</button>
    </div>
    <div class="legend">
      <span><span class="leg-dot" style="background:#d5f5e3;border:1px solid #aaa;"></span>Covered in Both</span>
      <span><span class="leg-dot" style="background:#fef9e7;border:1px solid #aaa;"></span>One panel only</span>
      <span><span class="leg-dot" style="background:#fadbd8;border:1px solid #aaa;"></span>NOT in Either</span>
      <span><span class="leg-dot" style="background:#f2f3f4;border:1px solid #aaa;"></span>Complex CNV</span>
      <span><span class="leg-dot" style="background:#fdf2f8;border:1px solid #aaa;"></span>TMS-Responsive</span>
    </div>
    <div id="all"   class="tab-pane active"><div class="tscroll">{full_table}</div></div>
    <div id="both"  class="tab-pane"><div class="tscroll">{both_table}</div></div>
    <div id="tonly" class="tab-pane"><div class="tscroll">{twist_only_table}</div></div>
    <div id="monly" class="tab-pane"><div class="tscroll">{medgeno_only_table}</div></div>
    <div id="ncov"  class="tab-pane"><div class="tscroll">{neither_table}</div></div>
    <div id="res"   class="tab-pane"><div class="tscroll">{res_table}</div></div>
  </div>

</div>
<footer>SNP Coverage Report — Twist SpikeinV2 &amp; Medgenome | Plasticity / TMS / Parkinson's Disease Genes</footer>

<script>
function showTab(id,btn){{
  document.querySelectorAll('.tab-pane').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById(id).classList.add('active'); btn.classList.add('active');
}}

new Chart(document.getElementById('overviewChart'),{{
  type:'bar',
  data:{{
    labels:['Both Covered','Twist Only','Medgenome Only','Neither','CNV'],
    datasets:[{{
      data:[{both_cov},{twist_only_count},{medgeno_only_count},{neither_count},{cnv}],
      backgroundColor:['#1E8449','#1F6AA5','#7D3C98','#C0392B','#7F8C8D']
    }}]
  }},
  options:{{responsive:true,plugins:{{legend:{{display:false}}}},
            scales:{{y:{{beginAtZero:true,ticks:{{stepSize:5}}}}}}}}
}});

new Chart(document.getElementById('geneChart'),{{
  type:'bar',
  data:{{
    labels:{gene_list},
    datasets:[
      {{label:'Twist',     data:{twist_cov_by_gene},   backgroundColor:'rgba(31,106,165,.75)'}},
      {{label:'Medgenome', data:{medgeno_cov_by_gene},  backgroundColor:'rgba(125,60,152,.75)'}}
    ]
  }},
  options:{{responsive:true,indexAxis:'y',plugins:{{legend:{{position:'bottom'}}}},
            scales:{{x:{{min:0,max:1,ticks:{{callback:v=>v==1?'Yes':'No'}}}}}}}}
}});

new Chart(document.getElementById('tmsChart'),{{
  type:'bar',
  data:{{
    labels:['Twist','Medgenome'],
    datasets:[
      {{label:'Covered',     data:[{tms_twist_cov},{tms_medgeno_cov}],     backgroundColor:['#1E8449','#7D3C98']}},
      {{label:'Not Covered', data:[{tms_total}-{tms_twist_cov},{tms_total}-{tms_medgeno_cov}], backgroundColor:['#C0392B','#E74C3C']}}
    ]
  }},
  options:{{responsive:true,plugins:{{legend:{{position:'bottom'}}}},
            scales:{{x:{{stacked:true}},y:{{stacked:true,max:{tms_total}}}}}}}
}});
</script>
</body>
</html>
"""
with open(OUT_HTML, "w") as fh:
    fh.write(html)
print(f"HTML  → {OUT_HTML}")
print(f"\n{'='*60}")
print(f"  Total SNP entries         : {total}")
print(f"  Covered in BOTH           : {both_cov}")
print(f"  Twist only                : {twist_only_count}")
print(f"  Medgenome only            : {medgeno_only_count}")
print(f"  NOT in either             : {neither_count}")
print(f"  Complex CNV               : {cnv}")
print(f"  rsIDs Resolved            : {resolved}")
print(f"  TMS Twist covered/total   : {tms_twist_cov}/{tms_total}")
print(f"  TMS Medgenome covered/tot : {tms_medgeno_cov}/{tms_total}")
